#!/usr/bin/env python3
"""
FerPs Anonymous – Anonymous submission bot with admin moderation + exports

Features
- Users DM the bot; their message is COPIED to @careerCampShadow anonymously.
- A copy goes to an admin group with buttons: 🗑 delete, 🚫 ban, 👤 profile, 🔗 view in channel.
- /user  → Admin-only list of unique senders (ChatID, @username, profile name).
- /info @username or /info <user_id> → Admin-only .xlsx export (timestamps, links, embedded photos/stickers).
- SQLite stores mappings, bans, and fields for exports.

Environment variables (or edit defaults below):
  BOT_TOKEN         – Telegram bot token (required)
  DEFAULT_CHANNEL   – Public @channel username (default: @ferpsanonymous)
  DB_PATH           – SQLite path (default: anonymous.db)
  ADMIN_IDS         – CSV of admin user IDs (default: "5821175466")

Dependencies:
  pip install "python-telegram-bot==20.8" "openpyxl==3.1.5" "Pillow==10.4.0"
  (Optional rate limiter) pip install 'python-telegram-bot[rate-limiter]==20.8'
"""
from __future__ import annotations

import html
import os
import sqlite3
import tempfile
from contextlib import closing
from pathlib import Path
from typing import Optional
import logging
from dotenv import load_dotenv
from pathlib import Path

# Load .env from the same folder as anon.py
load_dotenv(Path(__file__).resolve().parent / ".env")


logging.basicConfig(level=logging.INFO)
log = logging.getLogger("ferps_anon")   # or any name you want


from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Message, Update, constants
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

# Optional rate limiter (requires extra dependency). If missing, we run without it.
try:
    from telegram.ext import AIORateLimiter  # pip install 'python-telegram-bot[rate-limiter]==20.8'
except Exception:  # pragma: no cover
    AIORateLimiter = None

# --------------------------- Config ---------------------------------------
BOT_TOKEN = os.getenv("BOT_TOKEN")  # no default token here (safer)
DEFAULT_CHANNEL = os.getenv("DEFAULT_CHANNEL", "@ferpsanonymous")
DB_PATH = os.getenv("DB_PATH", "data/anonymous.db")
ADMIN_IDS = {int(x) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip()}

if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN is missing. Put it in .env")


# --------------------------- Storage --------------------------------------
SCHEMA = """
PRAGMA journal_mode=WAL;
CREATE TABLE IF NOT EXISTS settings (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS moderation (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    username TEXT,
    first_name TEXT,
    last_name TEXT,
    message_type TEXT NOT NULL,
    content_text TEXT,
    media_file_id TEXT,
    channel_username TEXT NOT NULL,
    channel_message_id INTEGER NOT NULL,
    group_message_id INTEGER NOT NULL,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS bans (
    user_id INTEGER PRIMARY KEY,
    reason TEXT,
    banned_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);
"""

def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db() -> None:
    with closing(db()) as conn:
        conn.executescript(SCHEMA)
        # seed default channel
        if conn.execute("SELECT value FROM settings WHERE key='CHANNEL_USERNAME'").fetchone() is None:
            conn.execute(
                "INSERT OR REPLACE INTO settings(key, value) VALUES('CHANNEL_USERNAME', ?)",
                (DEFAULT_CHANNEL,),
            )
        # migrations (older DBs)
        try:
            cols = {r[1] for r in conn.execute("PRAGMA table_info(moderation)").fetchall()}
            if "content_text" not in cols:
                conn.execute("ALTER TABLE moderation ADD COLUMN content_text TEXT")
            if "media_file_id" not in cols:
                conn.execute("ALTER TABLE moderation ADD COLUMN media_file_id TEXT")
        except Exception:
            pass
        conn.commit()

def get_setting(key: str) -> Optional[str]:
    with closing(db()) as conn:
        row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
        return row[0] if row else None

def set_setting(key: str, value: str) -> None:
    with closing(db()) as conn:
        conn.execute("INSERT OR REPLACE INTO settings(key, value) VALUES(?, ?)", (key, value))
        conn.commit()

# -------------------- Helpers -----------------------
def user_mention_html(user) -> str:
    name = html.escape((user.full_name or "Unknown").strip())
    return f"<a href='tg://user?id={user.id}'>{name}</a>"

def build_channel_link(channel_username: str, message_id: int) -> Optional[str]:
    if channel_username.startswith("@"):
        return f"https://t.me/{channel_username[1:]}/{message_id}"
    return None

async def is_group_admin(context: ContextTypes.DEFAULT_TYPE, chat_id: int, user_id: int) -> bool:
    try:
        member = await context.bot.get_chat_member(chat_id, user_id)
        return member.status in {"administrator", "creator"}
    except Exception:
        return False

async def requester_is_admin(user_id: int, context: ContextTypes.DEFAULT_TYPE) -> bool:
    if user_id in ADMIN_IDS:
        return True
    gid = get_setting("GROUP_CHAT_ID")
    if gid:
        try:
            member = await context.bot.get_chat_member(int(gid), user_id)
            return member.status in {"administrator", "creator"}
        except Exception:
            pass
    return False

# --------------------------- Commands -------------------------------------
async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.effective_message.reply_text(
        "👋 Welcome to FerPs Anonymous!\n\n"
        "Send me a message and I will post it anonymously to the channel."
)


async def cmd_setgroup(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    msg, chat = update.effective_message, update.effective_chat
    if chat.type not in ("group", "supergroup"):
        await msg.reply_text("Run this *inside* the admin group.", parse_mode=constants.ParseMode.MARKDOWN)
        return
    if not await is_group_admin(context, chat.id, update.effective_user.id):
        await msg.reply_text("Only group admins can set the group.")
        return
    set_setting("GROUP_CHAT_ID", str(chat.id))
    await msg.reply_text("✅ This chat is now registered as the admin group.")

async def cmd_setchannel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    msg, chat = update.effective_message, update.effective_chat
    if chat.type not in ("group", "supergroup"):
        await msg.reply_text("Run this inside the admin group.")
        return
    if not await is_group_admin(context, chat.id, update.effective_user.id):
        await msg.reply_text("Only group admins can set the channel.")
        return
    if not context.args:
        await msg.reply_text("Usage: /setchannel @ChannelUsername")
        return
    chan = context.args[0]
    if not chan.startswith("@"):
        await msg.reply_text("Please provide a public @channel username.")
        return
    set_setting("CHANNEL_USERNAME", chan)
    await msg.reply_text(f"✅ Channel set to {chan}")

async def cmd_stats(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    with closing(db()) as conn:
        total = conn.execute("SELECT COUNT(*) FROM moderation").fetchone()[0]
        banned = conn.execute("SELECT COUNT(*) FROM bans").fetchone()[0]
    await update.effective_message.reply_text(f"Total moderated posts: {total}\nBanned users: {banned}")

async def cmd_user(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Admin-only: list unique users who submitted (most-recent first)."""
    uid = update.effective_user.id
    if not await requester_is_admin(uid, context):
        await update.effective_message.reply_text("Admins only.")
        return
    with closing(db()) as conn:
        base = conn.execute(
            "SELECT user_id, MAX(created_at) AS last_at FROM moderation GROUP BY user_id ORDER BY last_at DESC"
        ).fetchall()
        if not base:
            await update.effective_message.reply_text("No users found yet.")
            return
        lines = ["Users who sent messages:"]
        for i, row in enumerate(base, start=1):
            user_id = row["user_id"]
            latest = conn.execute(
                "SELECT username, first_name, last_name FROM moderation WHERE user_id=? ORDER BY created_at DESC LIMIT 1",
                (user_id,),
            ).fetchone()
            username = (latest["username"] or "-")
            if username and not username.startswith("@") and username != "-":
                username = f"@{username}"
            full_name = (" ".join(filter(None, [latest["first_name"], latest["last_name"]])) or "-")
            lines.append(f"{i}. {user_id}(ChatID) - {username}(Username) - {full_name}(Profile name)")
    # split into safe chunks
    chunk = ""
    for line in lines:
        if len(chunk) + len(line) + 1 > 3500:
            await update.effective_message.reply_text(chunk)
            chunk = line
        else:
            chunk = (chunk + "\n" + line) if chunk else line
    if chunk:
        await update.effective_message.reply_text(chunk)

# --------------------------- /info export ---------------------------------
async def cmd_info(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Admin-only: /info @username OR /info <user_id> → XLSX with timestamps/links/photos."""
    req_id = update.effective_user.id
    if not await requester_is_admin(req_id, context):
        await update.effective_message.reply_text("Admins only.")
        return
    if not context.args:
        await update.effective_message.reply_text("Usage: /info @username or /info 123456789")
        return
    token = context.args[0].strip()
    if token.startswith("@"):
        token = token[1:]
    with closing(db()) as conn:
        if token.isdigit():
            rows = conn.execute(
                "SELECT * FROM moderation WHERE user_id=? ORDER BY created_at ASC", (int(token),)
            ).fetchall()
            label = token
        else:
            rows = conn.execute(
                "SELECT * FROM moderation WHERE username = ? COLLATE NOCASE ORDER BY created_at ASC", (token,)
            ).fetchall()
            label = f"@{token}"
    if not rows:
        await update.effective_message.reply_text("No submissions found for that user.")
        return
    path = await generate_xlsx(rows, label, context)
    try:
        await context.bot.send_document(
            chat_id=req_id, document=open(path, "rb"), filename=Path(path).name,
            caption=f"Submissions for {label} (total: {len(rows)})",
        )
    finally:
        try: os.remove(path)
        except Exception: pass

async def generate_xlsx(rows, user_label: str, context: ContextTypes.DEFAULT_TYPE) -> str:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage  # noqa: F401
    tmpdir = Path(tempfile.mkdtemp(prefix="ferps_anon_"))
    xlsx_path = tmpdir / f"info_{user_label.replace('@','')}.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Submissions"
    ws.append(["Message", "Sent (UTC)", "Type", "Channel Link", "Photo"])
    ws.column_dimensions['A'].width = 70
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 25
    row_idx = 2
    for r in rows:
        text = r["content_text"] or f"({r['message_type']})"
        when = str(r["created_at"])
        typ = r["message_type"]
        link = build_channel_link(r["channel_username"], r["channel_message_id"]) or ""
        ws.cell(row=row_idx, column=1, value=text)
        ws.cell(row=row_idx, column=2, value=when)
        ws.cell(row=row_idx, column=3, value=typ)
        if link:
            c = ws.cell(row=row_idx, column=4, value="Open")
            c.hyperlink = link; c.style = "Hyperlink"
        if typ in {"photo", "sticker"} and r["media_file_id"]:
            try:
                f = await context.bot.get_file(r["media_file_id"])
                ext = os.path.splitext(f.file_path or "")[1] or ".jpg"
                img_path = tmpdir / f"img_{row_idx}{ext}"
                await f.download_to_drive(img_path)
                ws.add_image(XLImage(str(img_path)), f"E{row_idx}")
                ws.row_dimensions[row_idx].height = 120
            except Exception:
                pass
        row_idx += 1
    wb.save(xlsx_path)
    return str(xlsx_path)

# --------------------------- Core logic -----------------------------------
async def handle_user_submission(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    msg, chat, user = update.effective_message, update.effective_chat, update.effective_user
    if chat.type != "private":
        return
    with closing(db()) as conn:
        if conn.execute("SELECT 1 FROM bans WHERE user_id=?", (user.id,)).fetchone():
            await msg.reply_text("🚫 You are banned from submitting messages.")
            return
    channel_username = get_setting("CHANNEL_USERNAME") or DEFAULT_CHANNEL
    group_id_str = get_setting("GROUP_CHAT_ID")
    if not group_id_str:
        await msg.reply_text("⚠️ The admin group isn't set yet. Ask an admin to run /setgroup in the group.")
        return
    group_id = int(group_id_str)
    try:
        channel_copy: Message = await msg.copy(chat_id=channel_username)
    except Exception as e:
        log.exception("Failed to copy to channel: %s", e)
        await msg.reply_text(f"❌ Couldn't post to channel.\nChannel: {channel_username}\nError: {e!r}")
        return

    kind = ("text" if msg.text else "photo" if msg.photo else "video" if msg.video else
            "voice" if msg.voice else "animation" if msg.animation else
            "sticker" if msg.sticker else "document" if msg.document else "other")
    content_text = msg.text or msg.caption
    media_file_id = None
    if msg.photo: media_file_id = msg.photo[-1].file_id
    elif msg.video: media_file_id = msg.video.file_id
    elif msg.voice: media_file_id = msg.voice.file_id
    elif msg.animation: media_file_id = msg.animation.file_id
    elif msg.sticker: media_file_id = msg.sticker.file_id
    elif msg.document: media_file_id = msg.document.file_id
    user_link = user_mention_html(user)
    submitted_text = html.escape(msg.text or msg.caption or "(no text)")
    header = (f"<b>New submission</b>\n"
              f"👤 {user_link}  (<code>{user.id}</code>)\n"
              f"🧾 Type: <code>{kind}</code>\n\n")
    body = f"<b>Message:</b>\n{submitted_text}"
    text_for_group = header + body
    try:
        if kind == "text":
            group_message = await context.bot.send_message(
                chat_id=group_id, text=text_for_group, parse_mode=constants.ParseMode.HTML
            )
        else:
            group_message = await msg.copy(
                chat_id=group_id, caption=text_for_group, parse_mode=constants.ParseMode.HTML
            )
    except Exception:
        log.exception("Failed to send to group")
        await msg.reply_text("I couldn't post to the admin group. Is the bot in that group?")
        return
    with closing(db()) as conn:
        cur = conn.execute(
            ("INSERT INTO moderation (user_id, username, first_name, last_name, message_type, content_text, media_file_id, "
             "channel_username, channel_message_id, group_message_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"),
            (user.id, user.username, user.first_name, user.last_name, kind,
             content_text, media_file_id, channel_username, channel_copy.message_id, group_message.message_id),
        )
        mod_id = cur.lastrowid
        conn.commit()
    chan_link = build_channel_link(channel_username, channel_copy.message_id)
    kb = [
        [InlineKeyboardButton("🗑", callback_data=f"del:{mod_id}"),
         InlineKeyboardButton("🚫", callback_data=f"ban:{mod_id}")],
        [InlineKeyboardButton("👤 Username", url=f"tg://user?id={user.id}"),
         InlineKeyboardButton("🔗 View in Channel", url=chan_link or "https://t.me/\u200b")],
    ]
    try:
        await context.bot.edit_message_reply_markup(
            chat_id=group_id, message_id=group_message.message_id, reply_markup=InlineKeyboardMarkup(kb)
        )
    except Exception:
        log.exception("Failed to attach keyboard to group message")
    await msg.reply_text("✅ Your message was sent anonymously to FerPs Anonymous.")


# --------------------------- Callbacks ------------------------------------
async def on_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    await q.answer()
    data = (q.data or "")
    action, _, id_part = data.partition(":")
    try:
        mod_id = int(id_part)
    except ValueError:
        return
    with closing(db()) as conn:
        row = conn.execute("SELECT * FROM moderation WHERE id=?", (mod_id,)).fetchone()
    if not row:
        await q.edit_message_reply_markup(reply_markup=None)
        return
    group_id = int(get_setting("GROUP_CHAT_ID") or 0)
    is_admin = (q.from_user.id in ADMIN_IDS)
    if not is_admin and q.message and q.message.chat and q.message.chat.id == group_id:
        is_admin = await is_group_admin(context, group_id, q.from_user.id)
    if not is_admin:
        await q.answer("Admins only.", show_alert=True)
        return
    channel_username, channel_msg_id, target_user_id = row["channel_username"], row["channel_message_id"], row["user_id"]
    if action == "del":
        try:
            await context.bot.delete_message(chat_id=channel_username, message_id=channel_msg_id)
            await q.answer("Deleted in channel")
        except Exception:
            await q.answer("Couldn't delete (maybe already deleted)")
        try:
            await q.edit_message_reply_markup(
                InlineKeyboardMarkup([[InlineKeyboardButton("🚫", callback_data=f"ban:{mod_id}")],
                                      [InlineKeyboardButton("👤 Username", url=f"tg://user?id={target_user_id}")]])
            )
        except Exception:
            pass
    elif action == "ban":
        with closing(db()) as conn:
            conn.execute("INSERT OR IGNORE INTO bans(user_id, reason) VALUES(?, ?)", (target_user_id, "Admin ban"))
            conn.commit()
        try:
            await context.bot.delete_message(chat_id=channel_username, message_id=channel_msg_id)
        except Exception:
            pass
        try:
            await context.bot.send_message(chat_id=target_user_id, text="🚫 You have been banned from submitting messages to this bot.")
        except Exception:
            pass
        await q.answer("User banned & post removed")
        try:
            await q.edit_message_reply_markup(reply_markup=None)
        except Exception:
            pass

# --------------------------- Misc handlers --------------------------------
async def handle_non_private(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    return

def build_app() -> Application:
    init_db()
    builder = Application.builder().token(BOT_TOKEN)
    if AIORateLimiter is not None:  # optional
        builder = builder.rate_limiter(AIORateLimiter())
    app = builder.build()
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("setgroup", cmd_setgroup))
    app.add_handler(CommandHandler("setchannel", cmd_setchannel))
    app.add_handler(CommandHandler("stats", cmd_stats))
    app.add_handler(CommandHandler("user", cmd_user))
    app.add_handler(CommandHandler("info", cmd_info))
    app.add_handler(CallbackQueryHandler(on_callback))
    app.add_handler(MessageHandler(filters.TEXT & filters.ChatType.PRIVATE & ~filters.COMMAND, handle_user_submission))
    # ✅ PTB v20 class-based media filters
    media_filters = (
        filters.PHOTO
        | filters.VIDEO
        | filters.VOICE
        | filters.ANIMATION
        | filters.Document.ALL
        | filters.Sticker.ALL
    )
    app.add_handler(MessageHandler(media_filters & filters.ChatType.PRIVATE, handle_user_submission))
    app.add_handler(MessageHandler(~filters.ChatType.PRIVATE, handle_non_private))
    return app

def main() -> None:
    """This bot now requires webhooks. Use main.py instead."""
    log.error("❌ This bot is webhook-only. Please run 'main.py' instead.")
    log.error("Example: python main.py")
    raise RuntimeError("Use main.py for webhook mode only")


if __name__ == "__main__":
    log.info("This bot is webhook-only. Use main.py instead.")
    main()
